#!/usr/bin/env python3
import os
import re
import mysql.connector
import pandas as pd
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 🔹 Configuration: Update your MySQL credentials
SQL_FILE_PATH = "/Users/example/Documents/db/db.sql"
OUTPUT_DIR = "/Users/example/Documents/db/output_excel"
MYSQL_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "root",  # 🔹 Change to your MySQL password
    "database": "temp_db"
}

# 🔹 Define tenant prefixes
TENANT_PREFIXES = [
    "blekastad", "brezina", "deml", "kalivoda", "komensky",
    "marci", "musil", "neumann", "pamatky", "patocka",
    "polanus", "sachs", "studenti", "tgm", "ucenci" # 🔹 etc
]

# 🔹 Create output directory
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 🔹 Function to load SQL file into MySQL
def load_sql_to_mysql(sql_file_path):
    connection = mysql.connector.connect(**MYSQL_CONFIG)
    cursor = connection.cursor()
    
    cursor.execute(f"CREATE DATABASE IF NOT EXISTS {MYSQL_CONFIG['database']};")
    cursor.execute(f"USE {MYSQL_CONFIG['database']};")

    with open(sql_file_path, "r", encoding="utf-8") as sql_file:
        sql_script = sql_file.read()

    statements = sql_script.split(";")
    for statement in statements:
        statement = statement.strip()
        if not statement:
            continue

        try:
            if statement.startswith("CREATE TABLE"):
                table_name = statement.split("`")[1]
                cursor.execute(f"SHOW TABLES LIKE '{table_name}';")
                if cursor.fetchone():
                    print(f"⚠️ Skipping existing table: {table_name}")
                    continue
            cursor.execute(statement)
        except mysql.connector.Error as err:
            print(f"⚠️ SQL Error: {err}")
            print(f"Problematic statement: {statement[:200]}...")  # Log problematic statement
            continue

    connection.commit()
    return connection

# 🔹 Function to expand JSON fields correctly
def expand_json_fields(df):
    """
    Expands JSON fields into separate columns.
    - JSON like {'cs': 'text', 'en': 'text'} → Becomes columns: `abstract/cs`, `abstract/en`.
    - JSON arrays like `[{"copy": "", "type": "letter"}]` → Becomes `copies/copy`, `copies/type`.
    """
    json_columns = [col for col in df.columns if df[col].apply(lambda x: isinstance(x, str) and x.strip().startswith(("{", "["))).any()]
    
    expanded_dfs = []  # List to store expanded DataFrames

    for col in json_columns:
        expanded_dict = {}

        for index, value in df[col].items():
            try:
                json_data = json.loads(value) if isinstance(value, str) else value
                if isinstance(json_data, dict):  # JSON object
                    for key, val in json_data.items():
                        expanded_dict.setdefault(f"{col}/{key}", {})[index] = val
                elif isinstance(json_data, list):  # JSON array (list of objects)
                    for obj in json_data:
                        if isinstance(obj, dict):
                            for key, val in obj.items():
                                expanded_dict.setdefault(f"{col}/{key}", {})[index] = val
            except json.JSONDecodeError:
                pass  # Ignore invalid JSON

        # Convert expanded columns into a DataFrame, aligned with original index
        expanded_df = pd.DataFrame.from_dict(expanded_dict, orient='index').T
        expanded_dfs.append(expanded_df)

    # Merge expanded columns with original DataFrame
    if expanded_dfs:
        df = pd.concat([df] + expanded_dfs, axis=1)
    
    # Drop original JSON columns after expansion
    df.drop(columns=json_columns, inplace=True)

    return df.copy()  # Avoid DataFrame fragmentation

# 🔹 Function to auto-adjust column widths in Excel
def adjust_excel_column_widths(writer, df, sheet_name):
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(df):
        series = df[col]
        max_len = max((
            series.astype(str).map(len).max(),  # Length of largest item
            len(str(col))  # Length of column name/header
        )) + 2  # Adding extra space
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len
        for cell in worksheet[get_column_letter(idx + 1)]:
            cell.alignment = Alignment(wrap_text=True)

# 🔹 Function to export MySQL tables to Excel
def export_tables_to_excel(connection):
    cursor = connection.cursor(dictionary=True)
    
    cursor.execute("SHOW TABLES;")
    tables = [table[f"Tables_in_{MYSQL_CONFIG['database']}"] for table in cursor.fetchall()]
    
    tenant_tables = {prefix: [] for prefix in TENANT_PREFIXES}
    global_tables = []

    for table in tables:
        matched = False
        for prefix in TENANT_PREFIXES:
            if table.startswith(prefix + "__"):
                tenant_tables[prefix].append(table)
                matched = True
                break
        if not matched:
            global_tables.append(table)

    for tenant, tables in tenant_tables.items():
        if not tables:
            continue
        output_file = os.path.join(OUTPUT_DIR, f"{tenant}.xlsx")
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for table in tables:
                cursor.execute(f"SELECT * FROM `{table}`;")
                rows = cursor.fetchall()
                if rows:
                    df = pd.DataFrame(rows)
                    df = expand_json_fields(df)  # Expand JSON fields
                    df.to_excel(writer, sheet_name=table.split("__", 1)[-1], index=False)
                    adjust_excel_column_widths(writer, df, table.split("__", 1)[-1])
        print(f"✅ Exported {tenant} -> {output_file}")

    if global_tables:
        global_output_file = os.path.join(OUTPUT_DIR, "global.xlsx")
        with pd.ExcelWriter(global_output_file, engine="openpyxl") as writer:
            for table in global_tables:
                cursor.execute(f"SELECT * FROM `{table}`;")
                rows = cursor.fetchall()
                if rows:
                    df = pd.DataFrame(rows)
                    df = expand_json_fields(df)  # Expand JSON fields
                    df.to_excel(writer, sheet_name=table, index=False)
                    adjust_excel_column_widths(writer, df, table)
        print(f"✅ Exported global tables -> {global_output_file}")

# 🔹 Main Execution
if __name__ == "__main__":
    connection = load_sql_to_mysql(SQL_FILE_PATH)
    export_tables_to_excel(connection)
    connection.close()
    print("🎉 Conversion completed successfully.")
